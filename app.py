from datetime import datetime
import os
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from flask import Flask, render_template, request, jsonify
from flask_cors import CORS
from chat import get_response

app = Flask(__name__)
#CORS(app)

@app.get("/")
def index_get():
    return render_template("base.html")

@app.post("/predict")
def predict():
    text = request.get_json().get("message")
    #Check if text is valid
    response = get_response(text)

    # Write the conversation to the Excel file
    #date_time = datetime.now()
    #if os.path.exists('Conversations.xlsx'):
    #    wb = load_workbook('Conversations.xlsx')
    #    sheet = wb.active
    #else:
    #    wb = Workbook()
    #    sheet = wb.active
    #    sheet.cell(row=1, column=1, value='Date')
    #    sheet.cell(row=1, column=2, value='User Message')
    #    sheet.cell(row=1, column=3, value='Bot Response')
    #sheet.insert_rows(2)
    #sheet.cell(row=2, column=1, value=date_time)
    #sheet.cell(row=2, column=2, value=text)
    #sheet.cell(row=2, column=3, value=response)
    #wb.save('Conversations.xlsx')

    message = {"answer": response}
    return message

if __name__=="__main__":
    app.run(debug=True)